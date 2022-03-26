#!/usr/bin/env python
# (c)2015 John Strickler

import openpyxl as px

COLUMN_HEADINGS = 'Name', 'Title', 'Favorite Color', 'Quest', 'Comment'

wb = px.Workbook()

ws = wb.active

ws.title = 'knights'

for column, heading in enumerate(COLUMN_HEADINGS, 1):
    ws.cell(row=1, column=column).value = heading
    ws.cell(row=1, column=column).font = px.styles.Font(bold=True)

with open('../DATA/knights.txt') as knights_in:
    for row, line in enumerate(knights_in, 2):
        name, title, color, quest, comment = line[:-1].split(':')
        ws.cell(row=row, column=1).value = name
        ws.cell(row=row, column=1).font = px.styles.Font(color='FFFF0000')
        ws.cell(row=row, column=2).value = title
        ws.cell(row=row, column=3).value = color
        ws.cell(row=row, column=4).value = quest
        ws.cell(row=row, column=5).value = comment
        ws.cell(row=row, column=5).font = px.styles.Font(italic=True)


wb.save('knights.xlsx')


