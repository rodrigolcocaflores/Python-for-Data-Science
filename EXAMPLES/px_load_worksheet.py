#!/usr/bin/env python
import openpyxl as px

def main():
    wb = px.load_workbook('../DATA/presidents.xlsx')

# three ways to get to a worksheet:

    # 1
    print(wb.sheetnames, '\n')
    ws = wb['US Presidents']
    print(ws, '\n')

    # 2
    for ws in wb:
        print(ws)
    print()

    # 3
    ws = wb.active
    print(ws, '\n')

    print(ws['A1'].value)
    print(ws['C2'].value, ws['B2'].value)


if __name__ == '__main__':
    main()
