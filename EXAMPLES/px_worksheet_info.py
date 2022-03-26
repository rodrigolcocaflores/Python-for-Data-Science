#!/usr/bin/env python
import openpyxl as px

def main():
    """program entry point"""
    wb = px.load_workbook('../DATA/presidents.xlsx')
    ws = wb['US Presidents']

    print(ws.dimensions)
    print(ws.min_column)
    print(ws.min_row)
    print(ws.max_column)
    print(ws.max_row)

    # returns value for specified cell
    print(ws.cell(row=2, column=3).value, ws.cell(row=2, column=2).value, '\n')

if __name__ == '__main__':
    main()
