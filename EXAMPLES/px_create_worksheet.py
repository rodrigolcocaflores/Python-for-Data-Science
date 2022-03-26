#!/usr/bin/env python
import openpyxl as px

fruits = ["pomegranate", "cherry", "apricot", "date", "apple", "lemon", "kiwi",
          "orange", "lime", "watermelon", "guava", "papaya", "fig", "pear", "banana",
          "tamarind", "persimmon", "elderberry", "peach", "blueberry", "lychee",
          "grape"]

wb = px.Workbook()

ws = wb.active

ws.title = 'fruits'

for i, fruit in enumerate(fruits, 1):
    ws.cell(row=i, column=1).value = fruit
    ws.cell(row=i, column=2).value = len(fruit)

wb.save('fruits.xlsx')
